"""
Recon spike, Part A: dictionary-file format census.

Walks O:/NaNDA/Data and catalogs every data-dictionary export found, so the Layer 2
parser can be written to normalize *every* variant in one pass. This is a FORMAT
CENSUS, not a presence check: it records filename patterns, file extensions, header
shapes, and which column holds the variable name.

Outputs (recon/out/):
  - dict_files.csv        one row per dictionary file found (path, classification, header)
  - dict_census.json      aggregate counts + header-variant catalog + gap lists
And prints a human-readable summary to stdout.

Run before building any generator. Read-only against O:.
"""

from __future__ import annotations

import csv
import json
import os
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

O_DATA = Path(r"O:\NaNDA\Data")
OUT_DIR = Path(__file__).resolve().parent / "out"

# Top-level entries under O:\NaNDA\Data that are not dataset topic folders.
NON_TOPIC = {"crosswalks", "z_archive", "_templates", "ZIPtoZCTA"}

# Path fragments that mark a dictionary file as archived / superseded / not the
# published export. Lowercased substring match against the full path.
EXCLUDE_PATH_FRAGMENTS = ("unused", "z_archive", "\\workfiles_received\\", "_archive", "~$")


def classify(path: Path) -> dict:
    """Classify a candidate dictionary file by extension, name pattern, and stage."""
    name = path.name.lower()
    parent = path.parent.name.lower()
    ext = path.suffix.lower()

    if "data_dictionary" in name:
        name_pattern = "*_data_dictionary" + ext
    elif "dictionary" in name:
        name_pattern = "*_dictionary" + ext
    else:
        name_pattern = "other" + ext

    if name.startswith("validate_"):
        stage = "validate"
    elif name.startswith("turnover_"):
        stage = "turnover"
    elif parent == "documentation":
        stage = "publish"  # publish-stage dict living in documentation\
    elif parent == "code":
        stage = "code"
    else:
        stage = "other"

    return {"ext": ext, "name_pattern": name_pattern, "stage": stage}


def read_header(path: Path) -> tuple[list[str] | None, str | None]:
    """Return (header_cells, error). Reads only the first row of the first sheet/line."""
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                for row in reader:
                    return [c.strip() for c in row], None
            return [], None  # empty file
        elif path.suffix.lower() in (".xlsx", ".xlsm"):
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() if c is not None else "" for c in row]
                break
            wb.close()
            return (header or []), None
    except Exception as exc:  # noqa: BLE001 - we want to record, not crash
        return None, f"{type(exc).__name__}: {exc}"
    return None, "unsupported extension"


def detect_varname_col(header: list[str]) -> str | None:
    """Which header cell names the variable column?"""
    if not header:
        return None
    lowered = [h.lower() for h in header]
    for candidate in ("variable", "name", "varname", "variable name", "field"):
        if candidate in lowered:
            return header[lowered.index(candidate)]
    return None


def is_candidate(path: Path) -> bool:
    name = path.name.lower()
    if path.suffix.lower() not in (".csv", ".xlsx", ".xlsm"):
        return False
    if "dictionary" not in name:
        return False
    full = str(path).lower()
    return not any(frag in full for frag in EXCLUDE_PATH_FRAGMENTS)


def main() -> int:
    if not O_DATA.exists():
        print(f"FATAL: {O_DATA} not reachable. Is the O: drive mapped?", file=sys.stderr)
        return 2

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    topic_folders = sorted(
        p for p in O_DATA.iterdir() if p.is_dir() and p.name not in NON_TOPIC
    )

    rows: list[dict] = []
    unparseable: list[dict] = []

    # Per dataset subfolder: does it have a published .dta, and does it have any dict?
    dataset_subfolders: dict[Path, dict] = {}

    for topic in topic_folders:
        for sub in sorted(p for p in topic.iterdir() if p.is_dir()):
            info = {"topic": topic.name, "subfolder": sub.name,
                    "has_P_dta": False, "has_any_dta": False, "dict_count": 0,
                    "csv_dict_count": 0, "xlsx_dict_count": 0, "publish_dict_count": 0}
            datasets_dir = sub / "datasets"
            if datasets_dir.is_dir():
                for f in datasets_dir.glob("*.dta"):
                    info["has_any_dta"] = True
                    if f.stem.lower().endswith("p"):
                        info["has_P_dta"] = True
            dataset_subfolders[sub] = info

    # Walk for dictionary files (documentation\ + code\, plus any nested).
    for topic in topic_folders:
        for path in topic.rglob("*"):
            if not path.is_file() or not is_candidate(path):
                continue
            cls = classify(path)
            header, err = read_header(path)
            varname_col = detect_varname_col(header) if header else None
            try:
                rel = path.relative_to(O_DATA)
            except ValueError:
                rel = path
            row = {
                "topic": path.relative_to(O_DATA).parts[0],
                "subfolder": path.relative_to(topic).parts[0],
                "filename": path.name,
                "ext": cls["ext"],
                "name_pattern": cls["name_pattern"],
                "stage": cls["stage"],
                "varname_col": varname_col or "",
                "ncols": len(header) if header else 0,
                "header": " | ".join(header) if header else "",
                "error": err or "",
                "relpath": str(rel),
            }
            rows.append(row)
            if err:
                unparseable.append({"relpath": str(rel), "error": err})

            # Credit dict to its dataset subfolder (walk up to the [shortname_daterange] dir).
            for anc in path.parents:
                if anc in dataset_subfolders:
                    dataset_subfolders[anc]["dict_count"] += 1
                    if cls["ext"] == ".csv":
                        dataset_subfolders[anc]["csv_dict_count"] += 1
                    elif cls["ext"] in (".xlsx", ".xlsm"):
                        dataset_subfolders[anc]["xlsx_dict_count"] += 1
                    if cls["stage"] == "publish":
                        dataset_subfolders[anc]["publish_dict_count"] += 1
                    break

    # --- Aggregates ---------------------------------------------------------
    by_ext = Counter(r["ext"] for r in rows)
    by_name_pattern = Counter(r["name_pattern"] for r in rows)
    by_stage = Counter(r["stage"] for r in rows)
    by_varname = Counter(r["varname_col"] or "(none)" for r in rows)

    # Distinct header shapes (only for publish-stage, the ones Layer 2 will read).
    header_variants: dict[str, dict] = defaultdict(lambda: {"count": 0, "examples": []})
    for r in rows:
        if r["stage"] != "publish" or not r["header"]:
            continue
        key = r["header"]
        hv = header_variants[key]
        hv["count"] += 1
        if len(hv["examples"]) < 3:
            hv["examples"].append(r["relpath"])

    # Dataset-subfolder gaps: published .dta present but no publish-stage dict.
    subs = list(dataset_subfolders.values())
    published_subs = [s for s in subs if s["has_P_dta"]]
    published_no_publish_dict = [s for s in published_subs if s["publish_dict_count"] == 0]
    published_only_xlsx = [
        s for s in published_subs
        if s["publish_dict_count"] > 0 and s["csv_dict_count"] == 0 and s["xlsx_dict_count"] > 0
    ]
    published_has_csv = [s for s in published_subs if s["csv_dict_count"] > 0]

    census = {
        "root": str(O_DATA),
        "topic_folder_count": len(topic_folders),
        "topic_folders": [t.name for t in topic_folders],
        "dataset_subfolder_count": len(subs),
        "dict_file_count": len(rows),
        "by_ext": dict(by_ext),
        "by_name_pattern": dict(by_name_pattern),
        "by_stage": dict(by_stage),
        "by_varname_col": dict(by_varname),
        "publish_header_variants": {
            k: v for k, v in sorted(header_variants.items(), key=lambda kv: -kv[1]["count"])
        },
        "published_subfolders_with_P_dta": len(published_subs),
        "published_subfolders_with_csv_dict": len(published_has_csv),
        "published_subfolders_only_xlsx_dict": [
            f'{s["topic"]}\\{s["subfolder"]}' for s in published_only_xlsx
        ],
        "published_subfolders_no_publish_dict": [
            f'{s["topic"]}\\{s["subfolder"]}' for s in published_no_publish_dict
        ],
        "unparseable": unparseable,
    }

    # --- Write artifacts ----------------------------------------------------
    with (OUT_DIR / "dict_files.csv").open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0].keys()) if rows else
                                ["topic", "subfolder", "filename", "ext", "name_pattern",
                                 "stage", "varname_col", "ncols", "header", "error", "relpath"])
        writer.writeheader()
        writer.writerows(rows)
    with (OUT_DIR / "dict_census.json").open("w", encoding="utf-8") as fh:
        json.dump(census, fh, indent=2)

    # --- Print summary ------------------------------------------------------
    print("=" * 70)
    print("DICTIONARY-FILE FORMAT CENSUS")
    print("=" * 70)
    print(f"Topic folders scanned        : {census['topic_folder_count']}")
    print(f"Dataset subfolders           : {census['dataset_subfolder_count']}")
    print(f"Dictionary files found       : {census['dict_file_count']}")
    print()
    print("By extension                 :", dict(by_ext))
    print("By filename pattern          :", dict(by_name_pattern))
    print("By pipeline stage            :", dict(by_stage))
    print("By variable-name column      :", dict(by_varname))
    print()
    print(f"Publish-stage header variants ({len(header_variants)} distinct):")
    for i, (hdr, v) in enumerate(sorted(header_variants.items(),
                                        key=lambda kv: -kv[1]["count"]), 1):
        print(f"  [{i}] x{v['count']}  cols=({hdr})")
        print(f"        e.g. {v['examples'][0]}")
    print()
    print(f"Published subfolders (have _P.dta)        : {len(published_subs)}")
    print(f"  ...with a CSV dictionary                : {len(published_has_csv)}")
    print(f"  ...with ONLY an xlsx dictionary         : {len(published_only_xlsx)}")
    for s in published_only_xlsx:
        print(f"        {s['topic']}\\{s['subfolder']}")
    print(f"  ...with NO publish-stage dictionary     : {len(published_no_publish_dict)}")
    for s in published_no_publish_dict:
        print(f"        {s['topic']}\\{s['subfolder']}")
    print()
    print(f"Unparseable dictionary files : {len(unparseable)}")
    for u in unparseable:
        print(f"        {u['relpath']}  ->  {u['error']}")
    print()
    print(f"Artifacts written to {OUT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
