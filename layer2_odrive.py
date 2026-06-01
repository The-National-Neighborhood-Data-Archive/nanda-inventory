"""
Layer 2 generator: O-drive file reality + internal mapping.

Source of truth is the O: drive itself. Scans O:\\NaNDA\\Data, and for every dataset
subfolder (one with published-looking data) reads the publish-stage dictionary exports in
its documentation\\ folder, then derives the internal-mapping fields from the variable list:

  geographies   - which census geographies the deposit is published at (+ boundary vintage)
  date_range    - from the folder's _[daterange] token, cross-checked with a `year` variable
  unique_ids    - the geographic identifier variable(s) (+ `year` for panels)
  smallest_geo  - the finest geography present

A subfolder can span several geographies (e.g. libraries: Tract10/20 + ZCTA10/20), so all of
its publish-stage dicts are parsed and unioned. CSV is preferred where a CSV dict exists;
otherwise xlsx (openpyxl). The variable-name column is auto-detected across the ~37 header
layouts the recon census found.

Selection of the *current* subfolder per catalog DOI, and the unique-ID diff against the v03
oracle, are the join's job (reconcile.py) — this layer just reports the raw O: reality, with
`has_P_dta` and `dict_status` so the join can apply the geo+date-range + _P tiebreaker.

Output: data/odrive_reality.csv (+ stdout summary). Read-only against O:.
"""

from __future__ import annotations

import csv
import re
import sys
from collections import Counter
from pathlib import Path

import openpyxl

O_DATA = Path(r"O:\NaNDA\Data")
OUT_CSV = Path(__file__).resolve().parent / "data" / "odrive_reality.csv"

NON_TOPIC = {"crosswalks", "z_archive", "_templates", "ZIPtoZCTA"}
EXCLUDE_PATH = ("unused", "z_archive", "workfile", "_archive", "~$",
                "data_dictionaries_unused")

# Geographic identifier variable patterns -> (geo label, rank). Lower rank = finer geography.
GEO_RANK = {"Block Group": 1, "Census Tract": 2, "ZCTA": 3, "ZIP": 3, "School District": 4,
            "County": 5, "State": 6}
GEO_PATTERNS = [
    ("Census Tract", re.compile(r"^tract_?fips(?P<v>10|20|22)?$")),
    ("Census Tract", re.compile(r"^trtid(?P<v>10|20)?$")),
    ("Census Tract", re.compile(r"^tract(?P<v>10|20)$")),
    ("Census Tract", re.compile(r"^geoid_?tract(?P<v>10|20)?$")),
    ("ZCTA",         re.compile(r"^zcta_?fips(?P<v>\d{2})?$")),
    ("ZCTA",         re.compile(r"^zcta(?P<v>\d{2,4})?$")),   # zcta, zcta10, zcta19, zcta20
    ("ZCTA",         re.compile(r"^zcta5$")),
    ("ZIP",          re.compile(r"^zip_?(?:code)?$")),
    ("Block Group",  re.compile(r"^bl[o]?ckgrp_?fips?(?P<v>10|20)?$")),
    ("Block Group",  re.compile(r"^bl[o]?ckgrp(?P<v>10|20)?$")),
    ("Block Group",  re.compile(r"^blkgrp.*$")),
    ("Block Group",  re.compile(r"^bg_?fips(?P<v>10|20)?$")),
    ("School District", re.compile(r"^leaid$")),
    ("School District", re.compile(r"^lea_?id$")),
    ("School District", re.compile(r"^district_?(id|fips)$")),
    ("County",       re.compile(r"^county_?fips$")),
    ("County",       re.compile(r"^countyfips$")),
    ("County",       re.compile(r"^stcofips(?P<v>\d{2})?$")),
    ("County",       re.compile(r"^fips$")),
    ("State",        re.compile(r"^state_?fips$")),
    ("State",        re.compile(r"^statefips$")),
]
YEAR_RE = re.compile(r"(?:19|20)\d{2}")
RANGE_RE = re.compile(r"((?:19|20)\d{2})\s*[-_]\s*((?:19|20)\d{2})")

# A dict filename encodes its file's publication geography (e.g. nanda_parks_Tract20_2024_
# dictionary.csv). This is the reliable signal for *publication* geography — far better than
# scanning variables, which also contain coarser rollup keys (e.g. stcofips inside a tract
# file). Order matters: check the finer/more-specific tokens first.
FILE_GEO_PATTERNS = [
    ("Block Group",     re.compile(r"bl[o]?ckgrp|blkgrp|blockgroup", re.I)),
    ("ZCTA",            re.compile(r"zcta(?P<v>10|20)?", re.I)),
    ("Census Tract",    re.compile(r"tract(?P<v>10|20)?", re.I)),
    ("School District", re.compile(r"district|leaid", re.I)),
    ("County",          re.compile(r"county", re.I)),
    ("ZIP",             re.compile(r"(?:^|[_-])zip(?:[_-]|$)", re.I)),
]


def geo_from_filename(name: str):
    """Publication geography + vintage inferred from a dict filename, else (None, None)."""
    for label, pat in FILE_GEO_PATTERNS:
        m = pat.search(name)
        if m:
            v = m.groupdict().get("v") if "v" in pat.groupindex else None
            return label, (f"20{v}" if v else None)
    return None, None


def detect_varname_col(header: list[str]) -> int | None:
    """Index of the variable-name column across the messy header layouts."""
    lowered = [(h or "").strip().lower() for h in header]
    for exact in ("variable", "variable name", "varname", "name", "field"):
        if exact in lowered:
            return lowered.index(exact)
    # e.g. "ZCTA10 Variable" / "Variable Name" variants -> first cell containing "variable"
    for i, h in enumerate(lowered):
        if "variable" in h:
            return i
    return 0 if header else None


def read_dict_columns(path: Path):
    """Return (variables:list[str], colmap:dict, error). colmap maps obs/min/max indices."""
    try:
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                rows = list(csv.reader(fh))
        else:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            rows = [[("" if c is None else str(c)) for c in r]
                    for r in ws.iter_rows(values_only=True)]
            wb.close()
        if not rows:
            return [], {}, None
        header = [str(c).strip() for c in rows[0]]
        vi = detect_varname_col(header)
        low = [h.lower() for h in header]
        colmap = {}
        for key, names in (("min", ("min", "minimum")), ("max", ("max", "maximum"))):
            for n in names:
                if n in low:
                    colmap[key] = low.index(n)
                    break
        variables, mins, maxs = [], [], []
        for r in rows[1:]:
            if vi is None or vi >= len(r):
                continue
            name = str(r[vi]).strip()
            if not name or name.lower() in ("variable", "nan"):
                continue
            variables.append(name)
            if "min" in colmap and colmap["min"] < len(r):
                mins.append(str(r[colmap["min"]]).strip())
            if "max" in colmap and colmap["max"] < len(r):
                maxs.append(str(r[colmap["max"]]).strip())
        return variables, {"mins": mins, "maxs": maxs, "vars": variables}, None
    except Exception as exc:  # noqa: BLE001
        return [], {}, f"{type(exc).__name__}: {exc}"


def classify_geo(var: str):
    """Return (geo_label, vintage|None) if var is a geographic identifier, else None."""
    v = var.strip().lower()
    for label, pat in GEO_PATTERNS:
        m = pat.match(v)
        if m:
            vintage = m.groupdict().get("v")
            return label, (f"20{vintage[-2:]}" if vintage else None)
    return None


def derive_from_year(colmap: dict, variables: list[str]) -> str | None:
    """If a `year` variable exists with min/max columns, return 'YYYY-YYYY'."""
    if "year" not in [v.lower() for v in variables]:
        return None
    idx = [v.lower() for v in variables].index("year")
    mins, maxs = colmap.get("mins", []), colmap.get("maxs", [])
    if idx < len(mins) and idx < len(maxs):
        ymin = YEAR_RE.search(mins[idx].replace(",", ""))
        ymax = YEAR_RE.search(maxs[idx].replace(",", ""))
        if ymin and ymax:
            return f"{ymin.group()}-{ymax.group()}" if ymin.group() != ymax.group() else ymin.group()
    return None


def date_from_folder(subfolder: str) -> str:
    m = RANGE_RE.search(subfolder)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    years = YEAR_RE.findall(subfolder)
    if years:
        return years[0] if len(set(years)) == 1 else f"{min(years)}-{max(years)}"
    return ""


def is_publish_dict(path: Path) -> bool:
    name = path.name.lower()
    if path.suffix.lower() not in (".csv", ".xlsx", ".xlsm"):
        return False
    if "dictionary" not in name:
        return False
    if name.startswith(("validate_", "turnover_")):
        return False
    if "documentation" not in [p.lower() for p in path.parts]:
        return False  # allow dicts nested under documentation\, not just its immediate child
    return not any(frag in str(path).lower() for frag in EXCLUDE_PATH)


def scan_subfolder(topic: str, sub: Path) -> dict | None:
    datasets = sub / "datasets"
    dtas = sorted(datasets.glob("*.dta")) if datasets.is_dir() else []
    if not dtas:
        return None  # not a dataset subfolder
    has_P = any(f.stem.lower().endswith("p") for f in dtas)

    docs = sub / "documentation"
    dict_files = sorted(p for p in docs.rglob("*") if p.is_file() and is_publish_dict(p)) \
        if docs.is_dir() else []

    geos: dict[str, set] = {}      # geo label -> set of vintages
    uid_vars: list[str] = []
    year_range = None
    n_vars_total = 0
    errors = []
    for df in dict_files:
        variables, colmap, err = read_dict_columns(df)
        if err:
            errors.append(f"{df.name}: {err}")
            continue
        n_vars_total += len(variables)

        # Publication geography for THIS file: filename first, else finest variable geo.
        pub_geo, file_vintage = geo_from_filename(df.name)
        if not pub_geo:
            var_labels = {classify_geo(v)[0] for v in variables if classify_geo(v)}
            pub_geo = min(var_labels, key=lambda g: GEO_RANK.get(g, 9)) if var_labels else None
        if not pub_geo:
            continue

        geos.setdefault(pub_geo, set())
        if file_vintage:
            geos[pub_geo].add(file_vintage)
        # Unique-ID vars = identifier variables matching the publication geography only
        # (excludes coarser rollup keys like stcofips inside a tract file).
        for v in variables:
            g = classify_geo(v)
            if g and g[0] == pub_geo:
                if g[1]:
                    geos[pub_geo].add(g[1])
                if v not in uid_vars:
                    uid_vars.append(v)
        if "year" in [v.lower() for v in variables] and "year" not in [u.lower() for u in uid_vars]:
            uid_vars.append("year")
        year_range = year_range or derive_from_year(colmap, variables)

    # Build geography labels with vintages, e.g. "Census Tract (2010, 2020); ZCTA (2020)".
    geo_label = "; ".join(
        f"{g}" + (f" ({', '.join(sorted(v))})" if v else "")
        for g, v in sorted(geos.items(), key=lambda kv: GEO_RANK.get(kv[0], 9))
    )
    smallest = min(geos, key=lambda g: GEO_RANK.get(g, 9)) if geos else ""
    folder_range = date_from_folder(sub.name)
    date_range = folder_range or (year_range or "")

    if not dict_files:
        dict_status = "NO_DICT"
    elif errors and not geos:
        dict_status = "PARSE_ERROR"
    else:
        dict_status = "ok"

    return {
        "topic_folder": topic,
        "subfolder": sub.name,
        "geographies": geo_label,
        "date_range": date_range,
        "date_range_folder": folder_range,
        "date_range_year_var": year_range or "",
        "unique_ids": "; ".join(uid_vars),
        "smallest_geo": smallest,
        "n_dicts": len(dict_files),
        "n_vars": n_vars_total,
        "has_P_dta": "Y" if has_P else "",
        "n_dta": len(dtas),
        "dict_status": dict_status,
        "parse_errors": " | ".join(errors),
    }


def main() -> int:
    if not O_DATA.exists():
        print(f"FATAL: {O_DATA} not reachable.", file=sys.stderr)
        return 2
    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)

    topics = sorted(p for p in O_DATA.iterdir() if p.is_dir() and p.name not in NON_TOPIC)
    rows = []
    for topic in topics:
        # A "dataset unit" is any directory holding a datasets\ or documentation\ child —
        # covers both layouts: [topic]\[shortname_daterange]\... (nested) and [topic]\...
        # (flat, e.g. ADI_standardized, air_conditioning, essential_businesses).
        units = set()
        for d in topic.rglob("*"):
            if not d.is_dir():
                continue
            if any(f in str(d).lower() for f in ("z_archive", "unused", "workfile", "_archive")):
                continue
            if d.name.lower() in ("datasets", "documentation"):
                units.add(d.parent)
        if (topic / "datasets").is_dir() or (topic / "documentation").is_dir():
            units.add(topic)
        for unit in sorted(units):
            row = scan_subfolder(topic.name, unit)
            if row:
                rows.append(row)

    cols = list(rows[0].keys()) if rows else []
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)

    no_geo = [r for r in rows if not r["geographies"]]
    no_dict = [r for r in rows if r["dict_status"] == "NO_DICT"]
    parse_err = [r for r in rows if r["dict_status"] == "PARSE_ERROR"]
    no_date = [r for r in rows if not r["date_range"]]
    by_smallest = Counter(r["smallest_geo"] or "(none)" for r in rows)

    print("=" * 70)
    print("LAYER 2 - O-DRIVE REALITY")
    print("=" * 70)
    print(f"Dataset subfolders (have .dta) : {len(rows)}  ->  {OUT_CSV}")
    print(f"  with derived geographies     : {len(rows) - len(no_geo)}")
    print(f"  with _P published .dta       : {sum(1 for r in rows if r['has_P_dta'])}")
    print(f"  smallest-geo breakdown       : {dict(by_smallest)}")
    print(f"  no dictionary on disk        : {len(no_dict)}")
    for r in no_dict:
        print(f"        {r['topic_folder']}\\{r['subfolder']}")
    print(f"  dictionary parse errors      : {len(parse_err)}")
    for r in parse_err:
        print(f"        {r['topic_folder']}\\{r['subfolder']}  {r['parse_errors'][:80]}")
    print(f"  rows with NO geography derived: {len(no_geo)}")
    for r in no_geo[:20]:
        print(f"        {r['topic_folder']}\\{r['subfolder']}  (dicts={r['n_dicts']}, status={r['dict_status']})")
    print(f"  rows with NO date_range      : {len(no_date)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
